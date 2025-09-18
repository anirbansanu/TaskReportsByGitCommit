#!/usr/bin/env python3
"""
Git Task Management Report Generator
Generates Excel reports from Git commit history across multiple repositories.
"""

import os
import json
import argparse
import subprocess
import sys
from datetime import datetime, timedelta
from collections import defaultdict, Counter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class GitTaskReportGenerator:
    def __init__(self):
        self.config = {}
        self.commits_data = []
        self.repo_stats = {}

    def parse_arguments(self):
        """Parse command line arguments"""
        parser = argparse.ArgumentParser(
            description='Generate task management Excel report from Git commit history'
        )
        parser.add_argument('--repos', nargs='+', help='Repository paths')
        parser.add_argument('--author', help='Filter commits by author')
        parser.add_argument('--since', help='Start date (YYYY-MM-DD)')
        parser.add_argument('--until', help='End date (YYYY-MM-DD)')
        parser.add_argument('--filename', help='Output Excel filename')
        parser.add_argument('--separate-sheets', action='store_true', 
                          help='Create separate sheet for each repository')

        args = parser.parse_args()

        # If no arguments provided, try to load from config file
        if not any(vars(args).values()):
            return self.load_config_file()

        return args

    def load_config_file(self, config_path='git_task_config.json'):
        """Load configuration from JSON file"""
        try:
            if not os.path.exists(config_path):
                # Create default config file
                default_config = {
                    "repos": ["./"],
                    "author": "",
                    "since": "2025-07-01",
                    "until": "2025-09-15",
                    "filename": "",
                    "separate_sheets": False
                }

                with open(config_path, 'w') as f:
                    json.dump(default_config, f, indent=2)

                print(f"Created default config file: {config_path}")
                print("Please edit the config file and run the script again.")
                sys.exit(0)

            with open(config_path, 'r') as f:
                config = json.load(f)

            # Convert to args-like object
            class ConfigArgs:
                def __init__(self, config):
                    self.repos = config.get('repos', ['./'])
                    self.author = config.get('author', '')
                    self.since = config.get('since', '')
                    self.until = config.get('until', '')
                    self.filename = config.get('filename', '')
                    setattr(self, 'separate_sheets', config.get('separate_sheets', False))

            print(f"Loaded configuration from {config_path}")
            return ConfigArgs(config)

        except Exception as e:
            print(f"Error loading config file: {e}")
            sys.exit(1)

    def validate_date(self, date_str, param_name):
        """Validate and parse date string"""
        if not date_str:
            return None

        try:
            return datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            print(f"Error: Invalid date format for {param_name}. Use YYYY-MM-DD format.")
            sys.exit(1)

    def validate_repository(self, repo_path):
        """Check if repository path exists and is a git repository"""
        if not os.path.exists(repo_path):
            print(f"Error: Repository path does not exist: {repo_path}")
            return False

        git_dir = os.path.join(repo_path, '.git')
        if not os.path.exists(git_dir):
            print(f"Error: Not a git repository: {repo_path}")
            return False

        return True

    def get_git_commits(self, repo_path, author=None, since=None, until=None):
        """Extract git commits from repository"""
        try:
            # Build git log command
            cmd = ['git', 'log', '--pretty=format:%H|%an|%ad|%s', '--date=short']

            if author:
                cmd.extend(['--author', author])

            if since:
                cmd.extend(['--since', since.strftime('%Y-%m-%d')])

            if until:
                cmd.extend(['--until', until.strftime('%Y-%m-%d')])

            # Execute git log command
            result = subprocess.run(
                cmd, 
                cwd=repo_path, 
                capture_output=True, 
                text=True,
                check=True
            )

            commits = []
            for line in result.stdout.strip().split('\n'):
                if not line:
                    continue

                parts = line.split('|', 3)
                if len(parts) >= 4:
                    commit_hash, commit_author, commit_date, commit_message = parts

                    # Skip merge commits
                    if commit_message.startswith('Merge'):
                        continue

                    commits.append({
                        'hash': commit_hash,
                        'author': commit_author,
                        'date': datetime.strptime(commit_date, '%Y-%m-%d').date(),
                        'message': commit_message.strip(),
                        'repo': os.path.basename(repo_path) if repo_path != './' else 'current'
                    })

            return commits

        except subprocess.CalledProcessError as e:
            print(f"Error executing git command in {repo_path}: {e}")
            return []
        except Exception as e:
            print(f"Error processing repository {repo_path}: {e}")
            return []

    def merge_commits_by_date(self, commits):
        """Merge commits that share the same date"""
        date_commits = defaultdict(list)
        
        # Group by date only (not by date AND repo)
        for commit in commits:
            date_commits[commit['date']].append(commit)
        
        merged_commits = []
        for commit_date, commits_list in date_commits.items():
            # Sort commits by repo name, then by message
            commits_list.sort(key=lambda x: (x['repo'], x['message']))
            
            # Format each commit message with repo name and join with line breaks
            formatted_messages = []
            for commit in commits_list:
                formatted_message = f"{commit['message']} ({commit['repo']})"
                formatted_messages.append(formatted_message)
            
            # Join messages with line breaks (ensure proper line break character)
            merged_message = '\n'.join(formatted_messages)
            
            merged_commits.append({
                'date': commit_date,
                'message': merged_message,
                'repo': 'combined' if len(set(c['repo'] for c in commits_list)) > 1 else commits_list[0]['repo'],
                'commit_count': len(commits_list)
            })
        
        return merged_commits



    def generate_task_rows(self, merged_commits):
        """Generate task management rows from merged commits"""
        tasks = []

        for commit in merged_commits:
            actual_end_date = commit['date']
            assign_date = actual_end_date - timedelta(days=1)
            planned_end_date = assign_date

            task = {
                'Task Name': commit['message'],
                'Task Priority': '',  # Leave blank
                'Assign Date': assign_date.strftime('%d-%m-%Y'),
                'Due Date': '',  # Leave blank
                'Planned End Date': planned_end_date.strftime('%d-%m-%Y'),
                'Actual End Date': actual_end_date.strftime('%d-%m-%Y'),
                'Assignee': 'Arvind Sir',
                'Repository': commit['repo'],
                'Commit Count': commit['commit_count']
            }

            tasks.append(task)

        # Sort by actual end date
        tasks.sort(key=lambda x: datetime.strptime(x['Actual End Date'], '%d-%m-%Y'))

        return tasks

    def generate_repo_statistics(self, all_commits):
        """Generate per-repository statistics"""
        stats = {}

        for repo in set(commit['repo'] for commit in all_commits):
            repo_commits = [c for c in all_commits if c['repo'] == repo]

            if not repo_commits:
                continue

            # Calculate statistics
            total_commits = len(repo_commits)
            date_counts = Counter(c['date'] for c in repo_commits)

            earliest_date = min(c['date'] for c in repo_commits)
            latest_date = max(c['date'] for c in repo_commits)

            date_range = (latest_date - earliest_date).days + 1
            avg_commits_per_day = total_commits / date_range if date_range > 0 else total_commits

            stats[repo] = {
                'Total Commits': total_commits,
                'Date Range': f"{earliest_date.strftime('%d-%m-%Y')} to {latest_date.strftime('%d-%m-%Y')}",
                'Days with Commits': len(date_counts),
                'Average Commits per Day': round(avg_commits_per_day, 2),
                'Max Commits in a Day': max(date_counts.values()) if date_counts else 0,
                'Authors': len(set(c['author'] for c in repo_commits))
            }

        return stats

    def create_excel_report(self, tasks, repo_stats, filename, separate_sheets=False):
        """Create Excel report with formatting"""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        if separate_sheets:
            # Create separate sheet for each repository
            repos = set(task['Repository'] for task in tasks)

            for repo in sorted(repos):
                repo_tasks = [task for task in tasks if task['Repository'] == repo]
                self.create_task_sheet(wb, f"Tasks_{repo}", repo_tasks)
        else:
            # Create single sheet for all repositories
            self.create_task_sheet(wb, "All_Tasks", tasks)

        # Create summary sheet
        self.create_summary_sheet(wb, repo_stats)

        # Save workbook
        wb.save(filename)
        print(f"Excel report saved: {filename}")

    def create_task_sheet(self, workbook, sheet_name, tasks):
        """Create a task sheet in the workbook"""
        ws = workbook.create_sheet(title=sheet_name)
        
        # Define columns (excluding Repository and Commit Count for display)
        columns = ['Task Name', 'Task Priority', 'Assign Date', 'Due Date', 
                'Planned End Date', 'Actual End Date', 'Assignee']
        
        # Add headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, task in enumerate(tasks, 2):
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=row, column=col, value=task[header])
                
                # Enable text wrapping for ALL Task Name cells
                if header == 'Task Name':
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    # Set row height based on content
                    if task[header] and '\n' in str(task[header]):
                        line_count = str(task[header]).count('\n') + 1
                        ws.row_dimensions[row].height = max(line_count * 15, 30)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        # For cells with line breaks, consider only the longest line
                        if '\n' in str(cell.value):
                            lines = str(cell.value).split('\n')
                            cell_length = max(len(line) for line in lines)
                        else:
                            cell_length = len(str(cell.value))
                        
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set wider width for Task Name column
            if column_letter == 'A':  # Task Name column
                adjusted_width = min(max_length + 5, 80)
            else:
                adjusted_width = min(max_length + 2, 50)
            
            ws.column_dimensions[column_letter].width = adjusted_width


    def create_summary_sheet(self, workbook, repo_stats):
        """Create summary sheet with repository statistics"""
        ws = workbook.create_sheet(title="Summary")

        row = 1

        # Title
        ws.cell(row=row, column=1, value="Repository Analysis Summary").font = Font(bold=True, size=14)
        row += 2

        for repo, stats in repo_stats.items():
            # Repository name
            ws.cell(row=row, column=1, value=f"Repository: {repo}").font = Font(bold=True)
            row += 1

            # Statistics
            for stat_name, stat_value in stats.items():
                ws.cell(row=row, column=2, value=stat_name)
                ws.cell(row=row, column=3, value=stat_value)
                row += 1

            row += 1  # Empty row between repositories

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_filename(self, since_date, until_date, custom_filename):
        """Generate output filename"""
        if custom_filename:
            return custom_filename

        if since_date and until_date:
            since_str = since_date.strftime('%Y%m%d')
            until_str = until_date.strftime('%Y%m%d')
            return f"task_report_{since_str}_to_{until_str}.xlsx"

        today = datetime.now().strftime('%Y%m%d')
        return f"task_report_{today}.xlsx"

    def run(self):
        """Main execution method"""
        try:
            # Parse arguments
            args = self.parse_arguments()

            # Validate dates
            since_date = self.validate_date(args.since, 'since')
            until_date = self.validate_date(args.until, 'until')

            # Validate repositories
            valid_repos = []
            for repo in args.repos:
                if self.validate_repository(repo):
                    valid_repos.append(repo)

            if not valid_repos:
                print("Error: No valid repositories found.")
                sys.exit(1)

            print(f"Processing {len(valid_repos)} repositories...")

            # Extract commits from all repositories
            all_commits = []

            for i, repo in enumerate(valid_repos, 1):
                print(f"[{i}/{len(valid_repos)}] Processing repository: {repo}")

                commits = self.get_git_commits(repo, args.author, since_date, until_date)
                all_commits.extend(commits)

                print(f"  Found {len(commits)} commits")

            if not all_commits:
                print("No commits found matching the criteria.")
                sys.exit(0)

            print(f"\nTotal commits found: {len(all_commits)}")

            # Generate repository statistics
            repo_stats = self.generate_repo_statistics(all_commits)

            # Generate tasks
            if getattr(args, 'separate_sheets', False):
                # Process each repository separately
                all_tasks = []

                for repo in set(commit['repo'] for commit in all_commits):
                    repo_commits = [c for c in all_commits if c['repo'] == repo]
                    merged_commits = self.merge_commits_by_date(repo_commits)
                    repo_tasks = self.generate_task_rows(merged_commits)
                    all_tasks.extend(repo_tasks)
            else:
                # Process all repositories together
                merged_commits = self.merge_commits_by_date(all_commits)
                all_tasks = self.generate_task_rows(merged_commits)

            # Generate filename
            filename = self.generate_filename(since_date, until_date, args.filename)

            # Create Excel report
            self.create_excel_report(
                all_tasks, 
                repo_stats, 
                filename, 
                getattr(args, 'separate_sheets', False)
            )

            print(f"\nTask report generated successfully!")
            print(f"Tasks created: {len(all_tasks)}")
            print(f"Output file: {filename}")

        except KeyboardInterrupt:
            print("\nOperation cancelled by user.")
            sys.exit(1)
        except Exception as e:
            print(f"Unexpected error: {e}")
            sys.exit(1)

if __name__ == "__main__":
    # Check required packages
    required_packages = ['pandas', 'openpyxl']
    missing_packages = []

    for package in required_packages:
        try:
            __import__(package)
        except ImportError:
            missing_packages.append(package)

    if missing_packages:
        print("Missing required packages. Please install:")
        for package in missing_packages:
            print(f"  pip install {package}")
        sys.exit(1)

    generator = GitTaskReportGenerator()
    generator.run()
